#!/usr/bin/env python3
"""
Stream one or more STAR BAMs and quantify nucleotide-conversion frequencies.

High-level workflow:
1. Read each BAM alignment record through `samtools view`.
2. Recover reference bases from an indexed FASTA and count ref->read conversions.
3. Bin each read by its internal conversion fraction.
4. Assign reads to genes using the supplied GTF feature type.
5. For each gene, compute:
   - fraction of reads above the per-read mismatch cutoff
   - absolute conversion fraction across all reads
   - absolute conversion fractions for above/below-cutoff subsets
6. Write either a single-dataset workbook or a combined multi-dataset workbook.

`argparse` provides `-h/--help` automatically; run the script with `--help`
to see the current CLI.
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from collections import Counter, defaultdict
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import DefaultDict, Dict, Iterator, List, Sequence, Tuple

from openpyxl import Workbook


CIGAR_OPS = set("MIDNSHP=X")
DEFAULT_SKIP_FLAG = 0x4 | 0x100 | 0x800
BUCKET_SIZE = 100_000
BASES = {"A", "C", "G", "T"}


@dataclass(frozen=True)
class FastaIndexEntry:
    length: int
    offset: int
    bases_per_line: int
    line_width: int


@dataclass(frozen=True)
class GeneInterval:
    start0: int
    end0: int
    gene: str


@dataclass
class GeneStats:
    assigned_reads: int = 0
    reads_above_threshold: int = 0
    reads_with_conversion: int = 0
    target_bases: int = 0
    conversions: int = 0
    target_bases_above_threshold: int = 0
    conversions_above_threshold: int = 0
    target_bases_below_threshold: int = 0
    conversions_below_threshold: int = 0
    target_bases_with_conversion: int = 0
    conversions_with_conversion: int = 0
    target_bases_without_conversion: int = 0
    conversions_without_conversion: int = 0


class IndexedFasta:
    def __init__(self, fasta_path: Path):
        self.fasta_path = fasta_path
        self.fai_path = fasta_path.with_suffix(fasta_path.suffix + ".fai")
        if not self.fai_path.exists():
            raise FileNotFoundError(
                f"Missing FASTA index: {self.fai_path}. Create it first with `samtools faidx`."
            )
        self.index = self._load_index(self.fai_path)
        self._fh = fasta_path.open("rb")
        self._cache_name = ""
        self._cache_start = 0
        self._cache_end = 0
        self._cache_seq = ""

    @staticmethod
    def _load_index(path: Path) -> Dict[str, FastaIndexEntry]:
        index: Dict[str, FastaIndexEntry] = {}
        with path.open("r", encoding="utf-8") as handle:
            for line in handle:
                fields = line.rstrip("\n").split("\t")
                if len(fields) < 5:
                    continue
                name = fields[0]
                index[name] = FastaIndexEntry(
                    length=int(fields[1]),
                    offset=int(fields[2]),
                    bases_per_line=int(fields[3]),
                    line_width=int(fields[4]),
                )
        return index

    def close(self) -> None:
        self._fh.close()

    def fetch(self, chrom: str, start0: int, end0: int) -> str:
        # Small on-demand cache keeps nearby genomic fetches cheap without
        # loading the full reference into memory.
        if start0 < 0 or end0 < start0:
            raise ValueError(f"Invalid FASTA range: {chrom}:{start0}-{end0}")
        if start0 == end0:
            return ""
        if (
            chrom == self._cache_name
            and start0 >= self._cache_start
            and end0 <= self._cache_end
        ):
            rel_start = start0 - self._cache_start
            rel_end = end0 - self._cache_start
            return self._cache_seq[rel_start:rel_end]

        entry = self.index.get(chrom)
        if entry is None:
            raise KeyError(f"Reference sequence `{chrom}` not found in {self.fasta_path}")
        if end0 > entry.length:
            raise ValueError(f"Requested {chrom}:{start0}-{end0} past contig length {entry.length}")

        seq = self._fetch_uncached(chrom, start0, end0, entry)
        span = end0 - start0
        if span < 1_000_000:
            cache_end = min(entry.length, max(end0, start0 + 1_000_000))
            self._cache_name = chrom
            self._cache_start = start0
            self._cache_end = cache_end
            self._cache_seq = self._fetch_uncached(chrom, start0, cache_end, entry)
        return seq

    def _fetch_uncached(
        self,
        chrom: str,
        start0: int,
        end0: int,
        entry: FastaIndexEntry,
    ) -> str:
        del chrom
        bases_per_line = entry.bases_per_line
        line_width = entry.line_width
        start_line = start0 // bases_per_line
        start_col = start0 % bases_per_line
        end_line = (end0 - 1) // bases_per_line
        end_col = (end0 - 1) % bases_per_line

        byte_start = entry.offset + start_line * line_width + start_col
        byte_end = entry.offset + end_line * line_width + end_col + 1

        self._fh.seek(byte_start)
        raw = self._fh.read(byte_end - byte_start)
        return raw.replace(b"\n", b"").replace(b"\r", b"").decode("ascii").upper()


def parse_cigar(cigar: str) -> Iterator[Tuple[int, str]]:
    number = []
    for char in cigar:
        if char.isdigit():
            number.append(char)
            continue
        if char not in CIGAR_OPS or not number:
            raise ValueError(f"Invalid CIGAR: {cigar}")
        yield int("".join(number)), char
        number.clear()
    if number:
        raise ValueError(f"Trailing length in CIGAR: {cigar}")


def aligned_blocks(cigar: str, ref_start0: int) -> Iterator[Tuple[int, int, int]]:
    """
    Yield tuples of (read_start0, ref_start0, length) for aligned query/reference blocks.
    Only M/=/X contribute aligned positions.
    """
    read_pos = 0
    ref_pos = ref_start0
    for length, op in parse_cigar(cigar):
        if op in {"M", "=", "X"}:
            yield read_pos, ref_pos, length
            read_pos += length
            ref_pos += length
        elif op in {"I", "S"}:
            read_pos += length
        elif op in {"D", "N"}:
            ref_pos += length
        elif op in {"H", "P"}:
            continue
        else:
            raise ValueError(f"Unsupported CIGAR op `{op}` in {cigar}")


def parse_gtf_attributes(attr_text: str) -> Dict[str, str]:
    attrs: Dict[str, str] = {}
    for part in attr_text.strip().split(";"):
        part = part.strip()
        if not part:
            continue
        key, _, value = part.partition(" ")
        if not key or not value:
            continue
        attrs[key] = value.strip().strip('"')
    return attrs


def choose_gene_name(attrs: Dict[str, str]) -> str | None:
    for key in ("gene_name", "gene_id", "transcript_id"):
        value = attrs.get(key)
        if value:
            return value
    return None


def load_gene_intervals(gtf_path: Path, feature: str) -> Dict[str, Dict[int, List[GeneInterval]]]:
    # Bucket intervals so overlap queries stay local to a genomic window.
    buckets: DefaultDict[str, DefaultDict[int, List[GeneInterval]]] = defaultdict(
        lambda: defaultdict(list)
    )
    with gtf_path.open("r", encoding="utf-8") as handle:
        for line in handle:
            if not line or line.startswith("#"):
                continue
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 9 or fields[2] != feature:
                continue
            chrom = fields[0]
            start0 = int(fields[3]) - 1
            end0 = int(fields[4])
            attrs = parse_gtf_attributes(fields[8])
            gene = choose_gene_name(attrs)
            if not gene:
                continue
            interval = GeneInterval(start0=start0, end0=end0, gene=gene)
            first_bucket = start0 // BUCKET_SIZE
            last_bucket = (end0 - 1) // BUCKET_SIZE
            for bucket_id in range(first_bucket, last_bucket + 1):
                buckets[chrom][bucket_id].append(interval)
    return {chrom: dict(bucket_map) for chrom, bucket_map in buckets.items()}


def overlapping_genes(
    interval_index: Dict[str, Dict[int, List[GeneInterval]]],
    chrom: str,
    start0: int,
    end0: int,
) -> set[str]:
    chrom_buckets = interval_index.get(chrom)
    if not chrom_buckets:
        return set()
    genes: set[str] = set()
    first_bucket = start0 // BUCKET_SIZE
    last_bucket = (end0 - 1) // BUCKET_SIZE
    for bucket_id in range(first_bucket, last_bucket + 1):
        for interval in chrom_buckets.get(bucket_id, ()):
            if interval.start0 < end0 and start0 < interval.end0:
                genes.add(interval.gene)
    return genes


def assign_gene(
    interval_index: Dict[str, Dict[int, List[GeneInterval]]],
    chrom: str,
    blocks: Sequence[Tuple[int, int, int]],
) -> str | None:
    genes: set[str] = set()
    for _, ref_start0, length in blocks:
        genes.update(overlapping_genes(interval_index, chrom, ref_start0, ref_start0 + length))
        if len(genes) > 1:
            return None
    return next(iter(genes)) if genes else None


def bin_index(fraction: float, bins: int) -> int:
    if fraction >= 1.0:
        return bins - 1
    return min(bins - 1, max(0, int(fraction * bins)))


def format_percent(value: float) -> str:
    return f"{value * 100:.2f}%"


def format_bin_label(index: int, bins: int) -> str:
    low = index / bins
    high = (index + 1) / bins
    return f"{low * 100:.1f}-{high * 100:.1f}%"


def build_samtools_command(
    samtools: str,
    bam_path: Path,
    min_mapq: int,
    skip_flag: int,
) -> List[str]:
    command = [samtools, "view", "-F", str(skip_flag)]
    if min_mapq > 0:
        command.extend(["-q", str(min_mapq)])
    command.append(str(bam_path))
    return command


def count_primary_mapped_reads(samtools: str, bam_path: Path, min_mapq: int, skip_flag: int) -> int:
    command = [samtools, "idxstats", str(bam_path)]
    process = subprocess.run(
        command,
        capture_output=True,
        text=True,
        encoding="utf-8",
        check=True,
    )
    total = 0
    for line in process.stdout.splitlines():
        fields = line.split("\t")
        if len(fields) >= 3:
            total += int(fields[2])
    return total


def emit_progress(
    reads_seen: int,
    total_reads_estimate: int | None,
    reads_with_target_base: int,
    total_conversions: int,
) -> None:
    message = f"[progress] reads={reads_seen:,} target_reads={reads_with_target_base:,} conversions={total_conversions:,}"
    if total_reads_estimate:
        pct = min(100.0, (reads_seen / total_reads_estimate) * 100.0)
        message += f" estimated_complete={pct:.2f}%"
    print(message, file=sys.stderr, flush=True)


def analyze_bam(
    bam_path: Path,
    fasta: IndexedFasta,
    ref_base: str,
    read_base: str,
    bins: int,
    gene_threshold: float,
    samtools: str,
    min_mapq: int,
    skip_flag: int,
    gene_intervals: Dict[str, Dict[int, List[GeneInterval]]] | None,
    max_reads: int | None,
    progress_interval: int,
    total_reads_estimate: int | None,
) -> Dict[str, object]:
    histogram = Counter()
    gene_stats: DefaultDict[str, GeneStats] = defaultdict(GeneStats)

    reads_seen = 0
    reads_with_target_base = 0
    total_target_bases = 0
    total_conversions = 0
    reads_with_any_conversion = 0
    reads_missing_reference = 0
    gene_assigned_reads = 0
    ambiguous_gene_reads = 0

    command = build_samtools_command(samtools, bam_path, min_mapq, skip_flag)
    process = subprocess.Popen(
        command,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        text=True,
        encoding="utf-8",
    )

    assert process.stdout is not None
    reached_limit = False
    try:
        for line in process.stdout:
            fields = line.rstrip("\n").split("\t")
            if len(fields) < 11:
                continue
            flag = int(fields[1])
            chrom = fields[2]
            pos1 = int(fields[3])
            cigar = fields[5]
            seq = fields[9].upper()

            if chrom == "*" or cigar == "*" or seq == "*":
                continue

            if max_reads is not None and reads_seen >= max_reads:
                reached_limit = True
                break
            reads_seen += 1
            blocks = list(aligned_blocks(cigar, pos1 - 1))
            if not blocks:
                continue

            read_target_bases = 0
            read_conversions = 0

            try:
                # Work directly from aligned query/reference blocks so STAR's
                # non-standard mismatch tags are irrelevant.
                for read_start0, ref_start0, length in blocks:
                    ref_seq = fasta.fetch(chrom, ref_start0, ref_start0 + length)
                    read_seq = seq[read_start0 : read_start0 + length]
                    for ref_nt, read_nt in zip(ref_seq, read_seq):
                        if ref_nt not in BASES or read_nt not in BASES:
                            continue
                        if ref_nt == ref_base:
                            read_target_bases += 1
                            if read_nt == read_base:
                                read_conversions += 1
            except KeyError:
                reads_missing_reference += 1
                continue

            if read_target_bases > 0:
                reads_with_target_base += 1
                total_target_bases += read_target_bases
                total_conversions += read_conversions
                if read_conversions > 0:
                    reads_with_any_conversion += 1
                fraction = read_conversions / read_target_bases
                histogram[bin_index(fraction, bins)] += 1

            if gene_intervals is None:
                continue

            gene = assign_gene(gene_intervals, chrom, blocks)
            if gene is None:
                ambiguous_gene_reads += 1
            else:
                gene_assigned_reads += 1
                stats = gene_stats[gene]
                stats.assigned_reads += 1
                stats.target_bases += read_target_bases
                stats.conversions += read_conversions

                if read_conversions > 0:
                    stats.reads_with_conversion += 1
                    stats.target_bases_with_conversion += read_target_bases
                    stats.conversions_with_conversion += read_conversions
                else:
                    stats.target_bases_without_conversion += read_target_bases
                    stats.conversions_without_conversion += read_conversions

                if read_target_bases > 0 and (read_conversions / read_target_bases) > gene_threshold:
                    stats.reads_above_threshold += 1
                    stats.target_bases_above_threshold += read_target_bases
                    stats.conversions_above_threshold += read_conversions
                else:
                    stats.target_bases_below_threshold += read_target_bases
                    stats.conversions_below_threshold += read_conversions

            if progress_interval > 0 and reads_seen % progress_interval == 0:
                emit_progress(
                    reads_seen=reads_seen,
                    total_reads_estimate=total_reads_estimate,
                    reads_with_target_base=reads_with_target_base,
                    total_conversions=total_conversions,
                )
    finally:
        if process.stdout is not None:
            process.stdout.close()
        if reached_limit and process.poll() is None:
            process.terminate()
            try:
                process.wait(timeout=2)
            except subprocess.TimeoutExpired:
                process.kill()
        stderr = process.stderr.read() if process.stderr is not None else ""
        return_code = process.wait()
        if return_code != 0 and not reached_limit:
            raise RuntimeError(f"`{' '.join(command)}` failed with code {return_code}:\n{stderr}")

    return {
        "reads_seen": reads_seen,
        "reads_with_target_base": reads_with_target_base,
        "reads_with_any_conversion": reads_with_any_conversion,
        "total_target_bases": total_target_bases,
        "total_conversions": total_conversions,
        "histogram": histogram,
        "gene_stats": gene_stats,
        "gene_assigned_reads": gene_assigned_reads,
        "ambiguous_gene_reads": ambiguous_gene_reads,
        "reads_missing_reference": reads_missing_reference,
    }


def default_output_path(bam_path: Path, ref_base: str, read_base: str) -> Path:
    return bam_path.with_name(f"{bam_path.stem}_{ref_base}2{read_base}_mismatch_analysis.xlsx")


def summarize_gene_threshold(
    gene_stats: Dict[str, GeneStats],
    threshold: float,
) -> Tuple[int, int, List[Tuple[str, GeneStats, float, float]]]:
    all_genes: List[Tuple[str, GeneStats, float, float]] = []
    reads_in_high_conversion_genes = 0
    total_assigned_reads = 0

    for gene, stats in gene_stats.items():
        total_assigned_reads += stats.assigned_reads
        if stats.assigned_reads == 0:
            continue
        pass_rate = stats.reads_above_threshold / stats.assigned_reads
        absolute_rate = (stats.conversions / stats.target_bases) if stats.target_bases else 0.0
        all_genes.append((gene, stats, pass_rate, absolute_rate))
        if pass_rate > threshold:
            reads_in_high_conversion_genes += stats.assigned_reads
    all_genes.sort(key=lambda item: (item[2], item[3]), reverse=True)

    return total_assigned_reads, reads_in_high_conversion_genes, all_genes


def summarize_single_mismatch_threshold(
    gene_stats: Dict[str, GeneStats],
) -> Tuple[int, int, List[Tuple[str, GeneStats, float, float]]]:
    all_genes: List[Tuple[str, GeneStats, float, float]] = []
    reads_in_positive_genes = 0
    total_assigned_reads = 0

    for gene, stats in gene_stats.items():
        total_assigned_reads += stats.assigned_reads
        if stats.assigned_reads == 0:
            continue
        pass_rate = stats.reads_with_conversion / stats.assigned_reads
        absolute_rate = (stats.conversions / stats.target_bases) if stats.target_bases else 0.0
        all_genes.append((gene, stats, pass_rate, absolute_rate))
        if stats.reads_with_conversion > 0:
            reads_in_positive_genes += stats.assigned_reads
    all_genes.sort(key=lambda item: (item[2], item[3]), reverse=True)

    return total_assigned_reads, reads_in_positive_genes, all_genes


def collect_summary_rows(results: Dict[str, object], threshold: float) -> Dict[str, object]:
    reads_seen = int(results["reads_seen"])
    reads_with_target_base = int(results["reads_with_target_base"])
    reads_with_any_conversion = int(results["reads_with_any_conversion"])
    total_target_bases = int(results["total_target_bases"])
    total_conversions = int(results["total_conversions"])
    gene_stats: Dict[str, GeneStats] = results["gene_stats"]  # type: ignore[assignment]
    gene_assigned_reads = int(results["gene_assigned_reads"])
    ambiguous_gene_reads = int(results["ambiguous_gene_reads"])
    reads_missing_reference = int(results["reads_missing_reference"])

    overall_rate = (total_conversions / total_target_bases) if total_target_bases else 0.0
    total_assigned_reads, high_reads, genes_over_threshold = summarize_gene_threshold(
        gene_stats, threshold
    )
    high_frac = (high_reads / total_assigned_reads) if total_assigned_reads else 0.0
    single_total_assigned_reads, single_high_reads, genes_single_mismatch = (
        summarize_single_mismatch_threshold(gene_stats)
    )
    single_high_frac = (
        (single_high_reads / single_total_assigned_reads) if single_total_assigned_reads else 0.0
    )

    return {
        "reads_seen": reads_seen,
        "reads_with_target_base": reads_with_target_base,
        "reads_with_any_conversion": reads_with_any_conversion,
        "total_target_bases": total_target_bases,
        "total_conversions": total_conversions,
        "overall_rate": overall_rate,
        "gene_assigned_reads": gene_assigned_reads,
        "ambiguous_gene_reads": ambiguous_gene_reads,
        "reads_missing_reference": reads_missing_reference,
        "total_assigned_reads": total_assigned_reads,
        "high_reads": high_reads,
        "high_frac": high_frac,
        "genes_over_threshold": genes_over_threshold,
        "single_high_reads": single_high_reads,
        "single_high_frac": single_high_frac,
        "genes_single_mismatch": genes_single_mismatch,
    }


def serialize_results(results: Dict[str, object]) -> Dict[str, object]:
    # Multi-dataset mode persists per-dataset summaries to temporary JSON files
    # before merging them into aligned workbook sheets.
    histogram: Counter = results["histogram"]  # type: ignore[assignment]
    gene_stats: Dict[str, GeneStats] = results["gene_stats"]  # type: ignore[assignment]
    return {
        "reads_seen": int(results["reads_seen"]),
        "reads_with_target_base": int(results["reads_with_target_base"]),
        "reads_with_any_conversion": int(results["reads_with_any_conversion"]),
        "total_target_bases": int(results["total_target_bases"]),
        "total_conversions": int(results["total_conversions"]),
        "gene_assigned_reads": int(results["gene_assigned_reads"]),
        "ambiguous_gene_reads": int(results["ambiguous_gene_reads"]),
        "reads_missing_reference": int(results["reads_missing_reference"]),
        "histogram": {str(k): int(v) for k, v in histogram.items()},
        "gene_stats": {
            gene: {
                "assigned_reads": stats.assigned_reads,
                "reads_above_threshold": stats.reads_above_threshold,
                "reads_with_conversion": stats.reads_with_conversion,
                "target_bases": stats.target_bases,
                "conversions": stats.conversions,
                "target_bases_above_threshold": stats.target_bases_above_threshold,
                "conversions_above_threshold": stats.conversions_above_threshold,
                "target_bases_below_threshold": stats.target_bases_below_threshold,
                "conversions_below_threshold": stats.conversions_below_threshold,
                "target_bases_with_conversion": stats.target_bases_with_conversion,
                "conversions_with_conversion": stats.conversions_with_conversion,
                "target_bases_without_conversion": stats.target_bases_without_conversion,
                "conversions_without_conversion": stats.conversions_without_conversion,
            }
            for gene, stats in gene_stats.items()
        },
    }


def deserialize_results(payload: Dict[str, object]) -> Dict[str, object]:
    return {
        "reads_seen": int(payload["reads_seen"]),
        "reads_with_target_base": int(payload["reads_with_target_base"]),
        "reads_with_any_conversion": int(payload["reads_with_any_conversion"]),
        "total_target_bases": int(payload["total_target_bases"]),
        "total_conversions": int(payload["total_conversions"]),
        "gene_assigned_reads": int(payload["gene_assigned_reads"]),
        "ambiguous_gene_reads": int(payload["ambiguous_gene_reads"]),
        "reads_missing_reference": int(payload["reads_missing_reference"]),
        "histogram": Counter({int(k): int(v) for k, v in dict(payload["histogram"]).items()}),
        "gene_stats": {
            gene: GeneStats(**stats) for gene, stats in dict(payload["gene_stats"]).items()
        },
    }


def autosize_worksheet(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [cell.value for cell in column_cells if cell.value is not None]
        if not values:
            continue
        width = max(len(str(value)) for value in values) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, 40)


def write_workbook(
    output_path: Path,
    bam_path: Path,
    ref_base: str,
    read_base: str,
    bins: int,
    threshold: float,
    results: Dict[str, object],
) -> None:
    histogram: Counter = results["histogram"]  # type: ignore[assignment]
    summary = collect_summary_rows(results, threshold)
    all_genes: List[Tuple[str, GeneStats, float, float]] = summary["genes_over_threshold"]  # type: ignore[assignment]
    single_genes: List[Tuple[str, GeneStats, float, float]] = summary["genes_single_mismatch"]  # type: ignore[assignment]

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Metric", "Value"])
    overview.append(["BAM", str(bam_path)])
    overview.append(["Reference base", ref_base])
    overview.append(["Read base", read_base])
    overview.append(["Reads analyzed", summary["reads_seen"]])
    overview.append(["Reads with target reference base", summary["reads_with_target_base"]])
    overview.append(["Reads with >=1 conversion", summary["reads_with_any_conversion"]])
    overview.append(["Target reference bases observed", summary["total_target_bases"]])
    overview.append(["Target conversions observed", summary["total_conversions"]])
    overview.append(["Overall conversion rate", summary["overall_rate"]])
    overview.append(["Reads skipped because contig absent from FASTA", summary["reads_missing_reference"]])

    per_read = workbook.create_sheet("Per_read_bins")
    per_read.append(
        ["Bin", "Reads", "Fraction_of_reads_with_target_base", "Percent_of_reads_with_target_base"]
    )
    reads_with_target_base = int(summary["reads_with_target_base"])
    for idx in range(bins):
        count = histogram[idx]
        fraction = (count / reads_with_target_base) if reads_with_target_base else 0.0
        per_read.append([format_bin_label(idx, bins), count, fraction, format_percent(fraction)])

    gene_threshold_sheet = workbook.create_sheet("Gene_threshold")
    gene_threshold_sheet.append(["Metric", "Value"])
    gene_threshold_sheet.append(["Reads assigned to exactly one gene", summary["gene_assigned_reads"]])
    gene_threshold_sheet.append(
        ["Reads skipped for gene summary because assignment was ambiguous or absent", summary["ambiguous_gene_reads"]]
    )
    gene_threshold_sheet.append(["Gene conversion threshold", threshold])
    gene_threshold_sheet.append(
        ["Genes reported", len(all_genes)]
    )
    gene_threshold_sheet.append(
        ["Reads in genes above threshold", summary["high_reads"]]
    )
    gene_threshold_sheet.append(
        ["Fraction of gene-assigned reads in genes above threshold", summary["high_frac"]]
    )
    gene_threshold_sheet.append(
        ["Percent of gene-assigned reads in genes above threshold", format_percent(float(summary["high_frac"]))]
    )
    gene_threshold_sheet.append(
        ["Reads in genes with >=1 mismatch-positive read", summary["single_high_reads"]]
    )
    gene_threshold_sheet.append(
        ["Fraction of gene-assigned reads in genes with >=1 mismatch-positive read", summary["single_high_frac"]]
    )

    genes_sheet = workbook.create_sheet("All_gene_pass_rates")
    genes_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_above_threshold",
            "Gene_read_pass_fraction",
            "Target_bases",
            "Conversions",
            "Absolute_conversion_fraction",
        ]
    )

    above_sheet = workbook.create_sheet("Above_threshold_rates")
    above_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_above_threshold",
            "Above_threshold_target_bases",
            "Above_threshold_conversions",
            "Above_threshold_absolute_fraction",
        ]
    )

    below_sheet = workbook.create_sheet("Below_threshold_rates")
    below_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_below_or_equal_threshold",
            "Below_or_equal_threshold_target_bases",
            "Below_or_equal_threshold_conversions",
            "Below_or_equal_threshold_absolute_fraction",
        ]
        )

    single_genes_sheet = workbook.create_sheet("Single_mismatch_pass_rates")
    single_genes_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_with_conversion",
            "Single_mismatch_pass_fraction",
            "Target_bases",
            "Conversions",
            "Absolute_conversion_fraction",
        ]
    )

    single_positive_sheet = workbook.create_sheet("Single_mismatch_positive_rates")
    single_positive_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_with_conversion",
            "Positive_target_bases",
            "Positive_conversions",
            "Positive_absolute_fraction",
        ]
    )

    single_zero_sheet = workbook.create_sheet("Single_mismatch_zero_rates")
    single_zero_sheet.append(
        [
            "Gene",
            "Assigned_reads",
            "Reads_without_conversion",
            "Zero_target_bases",
            "Zero_conversions",
            "Zero_absolute_fraction",
        ]
    )

    for gene, stats, pass_rate, absolute_rate in all_genes:
        above_absolute_rate = (
            stats.conversions_above_threshold / stats.target_bases_above_threshold
            if stats.target_bases_above_threshold
            else 0.0
        )
        below_absolute_rate = (
            stats.conversions_below_threshold / stats.target_bases_below_threshold
            if stats.target_bases_below_threshold
            else 0.0
        )
        genes_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.reads_above_threshold,
                pass_rate,
                stats.target_bases,
                stats.conversions,
                absolute_rate,
            ]
        )
        above_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.reads_above_threshold,
                stats.target_bases_above_threshold,
                stats.conversions_above_threshold,
                above_absolute_rate,
            ]
        )
        below_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.assigned_reads - stats.reads_above_threshold,
                stats.target_bases_below_threshold,
                stats.conversions_below_threshold,
                below_absolute_rate,
            ]
        )

    for gene, stats, pass_rate, absolute_rate in single_genes:
        positive_absolute_rate = (
            stats.conversions_with_conversion / stats.target_bases_with_conversion
            if stats.target_bases_with_conversion
            else 0.0
        )
        zero_absolute_rate = (
            stats.conversions_without_conversion / stats.target_bases_without_conversion
            if stats.target_bases_without_conversion
            else 0.0
        )
        single_genes_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.reads_with_conversion,
                pass_rate,
                stats.target_bases,
                stats.conversions,
                absolute_rate,
            ]
        )
        single_positive_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.reads_with_conversion,
                stats.target_bases_with_conversion,
                stats.conversions_with_conversion,
                positive_absolute_rate,
            ]
        )
        single_zero_sheet.append(
            [
                gene,
                stats.assigned_reads,
                stats.assigned_reads - stats.reads_with_conversion,
                stats.target_bases_without_conversion,
                stats.conversions_without_conversion,
                zero_absolute_rate,
            ]
        )

    for worksheet in workbook.worksheets:
        autosize_worksheet(worksheet)

    workbook.save(output_path)


def dataset_label(bam_path: Path) -> str:
    stem = bam_path.stem
    for suffix in ("_Aligned.sortedByCoord.out",):
        if stem.endswith(suffix):
            return stem[: -len(suffix)]
    return stem


def write_combined_workbook(
    output_path: Path,
    dataset_entries: List[Tuple[str, Path, Dict[str, object]]],
    ref_base: str,
    read_base: str,
    bins: int,
    threshold: float,
) -> None:
    summaries = [
        (label, bam_path, collect_summary_rows(results, threshold))
        for label, bam_path, results in dataset_entries
    ]
    all_genes = sorted(
        {
            gene
            for _, _, summary in summaries
            for gene, _, _, _ in summary["genes_over_threshold"]  # type: ignore[assignment]
        }
    )
    gene_metrics: Dict[str, Dict[str, Tuple[GeneStats, float, float]]] = {}
    single_gene_metrics: Dict[str, Dict[str, Tuple[GeneStats, float, float]]] = {}
    for label, _, summary in summaries:
        gene_metrics[label] = {
            gene: (stats, pass_rate, absolute_rate)
            for gene, stats, pass_rate, absolute_rate in summary["genes_over_threshold"]  # type: ignore[assignment]
        }
        single_gene_metrics[label] = {
            gene: (stats, pass_rate, absolute_rate)
            for gene, stats, pass_rate, absolute_rate in summary["genes_single_mismatch"]  # type: ignore[assignment]
        }

    workbook = Workbook()
    overview = workbook.active
    overview.title = "Overview"
    overview.append(["Metric"] + [label for label, _, _ in summaries])
    overview.append(["Reference base"] + [ref_base] * len(summaries))
    overview.append(["Read base"] + [read_base] * len(summaries))
    overview.append(["Gene threshold"] + [threshold] * len(summaries))
    overview.append(["BAM"] + [str(bam_path) for _, bam_path, _ in summaries])
    for metric_key, metric_label in (
        ("reads_seen", "Reads analyzed"),
        ("reads_with_target_base", "Reads with target reference base"),
        ("reads_with_any_conversion", "Reads with >=1 conversion"),
        ("total_target_bases", "Target reference bases observed"),
        ("total_conversions", "Target conversions observed"),
        ("overall_rate", "Overall conversion fraction"),
        ("gene_assigned_reads", "Reads assigned to exactly one gene"),
        ("ambiguous_gene_reads", "Reads skipped for gene summary"),
        ("reads_missing_reference", "Reads skipped because contig absent from FASTA"),
        ("high_frac", "Fraction of gene-assigned reads in genes above threshold"),
    ):
        overview.append([metric_label] + [summary[metric_key] for _, _, summary in summaries])

    per_read = workbook.create_sheet("Per_read_bins")
    per_read.append(["Bin"] + [label for label, _, _ in summaries])
    for idx in range(bins):
        row = [format_bin_label(idx, bins)]
        for (_, _, summary), (_, _, results) in zip(summaries, dataset_entries):
            reads_with_target_base = int(summary["reads_with_target_base"])
            histogram: Counter = results["histogram"]  # type: ignore[assignment]
            fraction = (histogram[idx] / reads_with_target_base) if reads_with_target_base else 0.0
            row.append(fraction)
        per_read.append(row)

    gene_threshold_sheet = workbook.create_sheet("Gene_threshold")
    gene_threshold_sheet.append(["Metric"] + [label for label, _, _ in summaries])
    for metric_key, metric_label in (
        ("gene_assigned_reads", "Reads assigned to exactly one gene"),
        ("ambiguous_gene_reads", "Reads skipped for gene summary because assignment was ambiguous or absent"),
        ("high_reads", "Reads in genes above threshold"),
        ("high_frac", "Fraction of gene-assigned reads in genes above threshold"),
        ("single_high_reads", "Reads in genes with >=1 mismatch-positive read"),
        ("single_high_frac", "Fraction of gene-assigned reads in genes with >=1 mismatch-positive read"),
    ):
        gene_threshold_sheet.append([metric_label] + [summary[metric_key] for _, _, summary in summaries])
    gene_threshold_sheet.append(["Genes reported"] + [len(summary["genes_over_threshold"]) for _, _, summary in summaries])

    genes_sheet = workbook.create_sheet("All_gene_pass_rates")
    genes_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_above_threshold",
                f"{label}_Gene_read_pass_fraction",
                f"{label}_Target_bases",
                f"{label}_Conversions",
                f"{label}_Absolute_conversion_fraction",
            )
        ]
    )

    above_sheet = workbook.create_sheet("Above_threshold_rates")
    above_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_above_threshold",
                f"{label}_Above_threshold_target_bases",
                f"{label}_Above_threshold_conversions",
                f"{label}_Above_threshold_absolute_fraction",
            )
        ]
    )

    below_sheet = workbook.create_sheet("Below_threshold_rates")
    below_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_below_or_equal_threshold",
                f"{label}_Below_or_equal_threshold_target_bases",
                f"{label}_Below_or_equal_threshold_conversions",
                f"{label}_Below_or_equal_threshold_absolute_fraction",
            )
        ]
    )

    single_genes_sheet = workbook.create_sheet("Single_mismatch_pass_rates")
    single_genes_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_with_conversion",
                f"{label}_Single_mismatch_pass_fraction",
                f"{label}_Target_bases",
                f"{label}_Conversions",
                f"{label}_Absolute_conversion_fraction",
            )
        ]
    )

    single_positive_sheet = workbook.create_sheet("Single_mismatch_positive_rates")
    single_positive_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_with_conversion",
                f"{label}_Positive_target_bases",
                f"{label}_Positive_conversions",
                f"{label}_Positive_absolute_fraction",
            )
        ]
    )

    single_zero_sheet = workbook.create_sheet("Single_mismatch_zero_rates")
    single_zero_sheet.append(
        ["Gene"] + [
            column
            for label, _, _ in summaries
            for column in (
                f"{label}_Assigned_reads",
                f"{label}_Reads_without_conversion",
                f"{label}_Zero_target_bases",
                f"{label}_Zero_conversions",
                f"{label}_Zero_absolute_fraction",
            )
        ]
    )

    for gene in all_genes:
        gene_row = [gene]
        above_row = [gene]
        below_row = [gene]
        single_gene_row = [gene]
        single_positive_row = [gene]
        single_zero_row = [gene]
        for label, _, _ in summaries:
            metrics = gene_metrics[label].get(gene)
            single_metrics = single_gene_metrics[label].get(gene)
            if metrics is None:
                gene_row.extend([0, 0, 0.0, 0, 0, 0.0])
                above_row.extend([0, 0, 0, 0, 0.0])
                below_row.extend([0, 0, 0, 0, 0.0])
            else:
                stats, pass_rate, absolute_rate = metrics
                above_absolute_rate = (
                    stats.conversions_above_threshold / stats.target_bases_above_threshold
                    if stats.target_bases_above_threshold
                    else 0.0
                )
                below_absolute_rate = (
                    stats.conversions_below_threshold / stats.target_bases_below_threshold
                    if stats.target_bases_below_threshold
                    else 0.0
                )
                gene_row.extend(
                    [
                        stats.assigned_reads,
                        stats.reads_above_threshold,
                        pass_rate,
                        stats.target_bases,
                        stats.conversions,
                        absolute_rate,
                    ]
                )
                above_row.extend(
                    [
                        stats.assigned_reads,
                        stats.reads_above_threshold,
                        stats.target_bases_above_threshold,
                        stats.conversions_above_threshold,
                        above_absolute_rate,
                    ]
                )
                below_row.extend(
                    [
                        stats.assigned_reads,
                        stats.assigned_reads - stats.reads_above_threshold,
                        stats.target_bases_below_threshold,
                        stats.conversions_below_threshold,
                        below_absolute_rate,
                    ]
                )

            if single_metrics is None:
                single_gene_row.extend([0, 0, 0.0, 0, 0, 0.0])
                single_positive_row.extend([0, 0, 0, 0, 0.0])
                single_zero_row.extend([0, 0, 0, 0, 0.0])
                continue

            single_stats, single_pass_rate, single_absolute_rate = single_metrics
            positive_absolute_rate = (
                single_stats.conversions_with_conversion / single_stats.target_bases_with_conversion
                if single_stats.target_bases_with_conversion
                else 0.0
            )
            zero_absolute_rate = (
                single_stats.conversions_without_conversion / single_stats.target_bases_without_conversion
                if single_stats.target_bases_without_conversion
                else 0.0
            )
            single_gene_row.extend(
                [
                    single_stats.assigned_reads,
                    single_stats.reads_with_conversion,
                    single_pass_rate,
                    single_stats.target_bases,
                    single_stats.conversions,
                    single_absolute_rate,
                ]
            )
            single_positive_row.extend(
                [
                    single_stats.assigned_reads,
                    single_stats.reads_with_conversion,
                    single_stats.target_bases_with_conversion,
                    single_stats.conversions_with_conversion,
                    positive_absolute_rate,
                ]
            )
            single_zero_row.extend(
                [
                    single_stats.assigned_reads,
                    single_stats.assigned_reads - single_stats.reads_with_conversion,
                    single_stats.target_bases_without_conversion,
                    single_stats.conversions_without_conversion,
                    zero_absolute_rate,
                ]
            )
        genes_sheet.append(gene_row)
        above_sheet.append(above_row)
        below_sheet.append(below_row)
        single_genes_sheet.append(single_gene_row)
        single_positive_sheet.append(single_positive_row)
        single_zero_sheet.append(single_zero_row)

    for worksheet in workbook.worksheets:
        autosize_worksheet(worksheet)

    workbook.save(output_path)


def analyze_one_dataset(
    bam_path: Path,
    reference_fasta: Path,
    gtf: Path | None,
    gene_feature: str,
    ref_base: str,
    read_base: str,
    bins: int,
    gene_threshold: float,
    samtools: str,
    min_mapq: int,
    max_reads: int | None,
    progress_interval: int,
) -> Dict[str, object]:
    gene_intervals = load_gene_intervals(gtf, gene_feature) if gtf is not None else None
    try:
        total_reads_estimate = count_primary_mapped_reads(
            samtools,
            bam_path,
            min_mapq,
            DEFAULT_SKIP_FLAG,
        )
    except subprocess.CalledProcessError:
        total_reads_estimate = None

    fasta = IndexedFasta(reference_fasta)
    try:
        return analyze_bam(
            bam_path=bam_path,
            fasta=fasta,
            ref_base=ref_base,
            read_base=read_base,
            bins=bins,
            gene_threshold=gene_threshold,
            samtools=samtools,
            min_mapq=min_mapq,
            skip_flag=DEFAULT_SKIP_FLAG,
            gene_intervals=gene_intervals,
            max_reads=max_reads,
            progress_interval=progress_interval,
            total_reads_estimate=total_reads_estimate,
        )
    finally:
        fasta.close()


def analyze_dataset_to_temp(
    bam_path_str: str,
    reference_fasta_str: str,
    gtf_str: str | None,
    gene_feature: str,
    ref_base: str,
    read_base: str,
    bins: int,
    gene_threshold: float,
    samtools: str,
    min_mapq: int,
    max_reads: int | None,
    progress_interval: int,
    temp_dir_str: str,
) -> Tuple[str, str, str]:
    bam_path = Path(bam_path_str)
    results = analyze_one_dataset(
        bam_path=bam_path,
        reference_fasta=Path(reference_fasta_str),
        gtf=Path(gtf_str) if gtf_str is not None else None,
        gene_feature=gene_feature,
        ref_base=ref_base,
        read_base=read_base,
        bins=bins,
        gene_threshold=gene_threshold,
        samtools=samtools,
        min_mapq=min_mapq,
        max_reads=max_reads,
        progress_interval=progress_interval,
    )
    label = dataset_label(bam_path)
    temp_path = Path(temp_dir_str) / f"{label}.json"
    temp_path.write_text(json.dumps(serialize_results(results)), encoding="utf-8")
    return label, str(bam_path), str(temp_path)



def print_summary(results: Dict[str, object], bins: int, threshold: float) -> None:
    histogram: Counter = results["histogram"]  # type: ignore[assignment]
    summary = collect_summary_rows(results, threshold)
    reads_seen = int(summary["reads_seen"])
    reads_with_target_base = int(summary["reads_with_target_base"])

    print("Overall")
    print(f"Reads analyzed: {reads_seen}")
    print(f"Reads with target reference base: {summary['reads_with_target_base']}")
    print(f"Reads with >=1 conversion: {summary['reads_with_any_conversion']}")
    print(f"Target reference bases observed: {summary['total_target_bases']}")
    print(f"Target conversions observed: {summary['total_conversions']}")
    print(f"Overall conversion rate: {format_percent(float(summary['overall_rate']))}")
    if int(summary["reads_missing_reference"]):
        print(
            "Reads skipped because contig was absent from FASTA: "
            f"{summary['reads_missing_reference']}"
        )
    print()

    print("Per-read bins")
    print("Bin\tReads\tPercent_of_reads_with_target_base")
    for idx in range(bins):
        count = histogram[idx]
        frac = (count / reads_with_target_base) if reads_with_target_base else 0.0
        print(f"{format_bin_label(idx, bins)}\t{count}\t{format_percent(frac)}")
    print()

    all_genes: List[Tuple[str, GeneStats, float, float]] = summary["genes_over_threshold"]  # type: ignore[assignment]
    if not results["gene_stats"]:
        print("Gene threshold")
        print("No gene annotation was provided; gene-level statistics were not calculated.")
        return

    print("Gene threshold")
    print(f"Reads assigned to exactly one gene: {summary['gene_assigned_reads']}")
    print(
        "Reads skipped for gene summary because assignment was ambiguous or absent: "
        f"{summary['ambiguous_gene_reads']}"
    )
    print(f"Genes reported: {len(all_genes)}")
    print(
        "Percent of gene-assigned reads mapping to genes above threshold: "
        f"{format_percent(float(summary['high_frac']))}"
    )
    print()

    if all_genes:
        print("Top genes by gene-level read-pass percent")
        print(
            "Gene\tAssigned_reads\tReads_above_threshold\tGene_read_pass_percent\t"
            "Absolute_conversion_percent"
        )
        for gene, stats, pass_rate, absolute_rate in all_genes[:20]:
            print(
                f"{gene}\t{stats.assigned_reads}\t{stats.reads_above_threshold}\t"
                f"{format_percent(pass_rate)}\t{format_percent(absolute_rate)}"
            )


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Quantify nucleotide-conversion frequencies from one or more STAR BAMs. "
            "Each primary alignment record is treated as one read; run with --help "
            "to see all options."
        ),
        epilog=(
            "Example single dataset: RNAseq_mismatch_analysis.py sample.bam "
            "--reference-fasta ref.fna --gtf genomic.gtf --ref-base C --read-base G\n"
            "Example combined datasets: RNAseq_mismatch_analysis.py a.bam b.bam "
            "--reference-fasta ref.fna --gtf genomic.gtf --jobs 2"
        ),
    )
    parser.add_argument("bam", nargs="+", type=Path, help="One or more input BAM files")
    parser.add_argument(
        "--reference-fasta",
        required=True,
        type=Path,
        help="Reference FASTA matching the BAM contig names; requires an existing .fai index",
    )
    parser.add_argument(
        "--gtf",
        type=Path,
        help="Optional gene annotation GTF used to assign reads to genes by exon overlap",
    )
    parser.add_argument(
        "--gene-feature",
        default="exon",
        help="GTF feature type used for gene assignment [default: exon]",
    )
    parser.add_argument("--ref-base", default="C", choices=sorted(BASES))
    parser.add_argument("--read-base", default="G", choices=sorted(BASES))
    parser.add_argument(
        "--bins",
        type=int,
        default=10,
        help="Number of per-read conversion bins across 0-100%% [default: 10]",
    )
    parser.add_argument(
        "--gene-threshold",
        type=float,
        default=0.10,
        help="Gene-level conversion-rate threshold used for the final read percentage [default: 0.10]",
    )
    parser.add_argument(
        "--min-mapq",
        type=int,
        default=0,
        help="Minimum MAPQ passed to samtools view [default: 0]",
    )
    parser.add_argument(
        "--samtools",
        default="samtools",
        help="Path to samtools executable [default: samtools]",
    )
    parser.add_argument(
        "--output",
        type=Path,
        help="Output Excel workbook path [.xlsx]. Defaults to BAM stem plus conversion label.",
    )
    parser.add_argument(
        "--max-reads",
        type=int,
        help="Stop after analyzing this many mapped primary alignments.",
    )
    parser.add_argument(
        "--progress-interval",
        type=int,
        default=100000,
        help="Emit a progress line to stderr every N analyzed reads [default: 100000]",
    )
    parser.add_argument(
        "--jobs",
        type=int,
        help="Number of datasets to analyze in parallel when multiple BAMs are provided",
    )
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()

    if args.ref_base == args.read_base:
        parser.error("--ref-base and --read-base must differ.")
    if args.bins <= 0:
        parser.error("--bins must be a positive integer.")
    if not (0.0 <= args.gene_threshold <= 1.0):
        parser.error("--gene-threshold must be between 0 and 1.")
    if args.max_reads is not None and args.max_reads <= 0:
        parser.error("--max-reads must be a positive integer.")
    if args.progress_interval <= 0:
        parser.error("--progress-interval must be a positive integer.")
    if args.jobs is not None and args.jobs <= 0:
        parser.error("--jobs must be a positive integer.")

    if len(args.bam) == 1:
        bam_path = args.bam[0]
        results = analyze_one_dataset(
            bam_path=bam_path,
            reference_fasta=args.reference_fasta,
            gtf=args.gtf,
            gene_feature=args.gene_feature,
            ref_base=args.ref_base,
            read_base=args.read_base,
            bins=args.bins,
            gene_threshold=args.gene_threshold,
            samtools=args.samtools,
            min_mapq=args.min_mapq,
            max_reads=args.max_reads,
            progress_interval=args.progress_interval,
        )

        output_path = args.output or default_output_path(bam_path, args.ref_base, args.read_base)
        write_workbook(
            output_path=output_path,
            bam_path=bam_path,
            ref_base=args.ref_base,
            read_base=args.read_base,
            bins=args.bins,
            threshold=args.gene_threshold,
            results=results,
        )
        print_summary(results, bins=args.bins, threshold=args.gene_threshold)
        print()
        print(f"Workbook written: {output_path}")
        return 0

    jobs = args.jobs or min(len(args.bam), os.cpu_count() or len(args.bam))
    output_path = args.output or (Path.cwd() / f"combined_{args.ref_base}2{args.read_base}_mismatch_analysis.xlsx")
    dataset_entries: List[Tuple[str, Path, Dict[str, object]]] = []

    with TemporaryDirectory(prefix="rnaseq_mismatch_") as temp_dir:
        try:
            executor_factory = ProcessPoolExecutor
            executor_context = executor_factory(max_workers=jobs)
        except PermissionError:
            executor_factory = ThreadPoolExecutor
            executor_context = executor_factory(max_workers=jobs)

        with executor_context as executor:
            futures = [
                executor.submit(
                    analyze_dataset_to_temp,
                    str(bam_path),
                    str(args.reference_fasta),
                    str(args.gtf) if args.gtf is not None else None,
                    args.gene_feature,
                    args.ref_base,
                    args.read_base,
                    args.bins,
                    args.gene_threshold,
                    args.samtools,
                    args.min_mapq,
                    args.max_reads,
                    args.progress_interval,
                    temp_dir,
                )
                for bam_path in args.bam
            ]
            for future in as_completed(futures):
                label, bam_path_str, temp_json_str = future.result()
                payload = json.loads(Path(temp_json_str).read_text(encoding="utf-8"))
                dataset_entries.append((label, Path(bam_path_str), deserialize_results(payload)))

    dataset_entries.sort(key=lambda item: item[0])
    write_combined_workbook(
        output_path=output_path,
        dataset_entries=dataset_entries,
        ref_base=args.ref_base,
        read_base=args.read_base,
        bins=args.bins,
        threshold=args.gene_threshold,
    )
    print(f"Combined workbook written: {output_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
